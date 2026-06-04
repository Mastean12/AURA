import os
import re
from pathlib import Path

ALLOWED_EXTENSIONS = {".pdf", ".docx", ".csv", ".xlsx"}


def allowed_file(filename: str) -> bool:
    ext = Path(filename).suffix.lower()
    return ext in ALLOWED_EXTENSIONS


def clean_text(text: str) -> str:
    text = re.sub(r"\r\n", "\n", text)
    text = re.sub(r"\r", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_text(filepath: str) -> str:
    ext = Path(filepath).suffix.lower()

    if ext == ".pdf":
        raw = _extract_pdf(filepath)
    elif ext == ".docx":
        raw = _extract_docx(filepath)
    elif ext == ".csv":
        raw = _extract_csv(filepath)
    elif ext == ".xlsx":
        raw = _extract_xlsx(filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return clean_text(raw)


def _extract_pdf(filepath: str) -> str:
    from PyPDF2 import PdfReader

    reader = PdfReader(filepath)
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    return "\n\n".join(pages)


def _extract_docx(filepath: str) -> str:
    from docx import Document

    doc = Document(filepath)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def _extract_csv(filepath: str) -> str:
    import pandas as pd

    df = pd.read_csv(filepath)
    lines = []
    lines.append(",".join(str(c) for c in df.columns))
    for _, row in df.iterrows():
        lines.append(",".join(str(v) if pd.notna(v) else "" for v in row))
    return "\n".join(lines)


def _extract_xlsx(filepath: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            row_text = ",".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip(","):
                lines.append(row_text)
    wb.close()
    return "\n".join(lines)


def save_upload(content: bytes, filename: str, upload_dir: str) -> str:
    filepath = os.path.join(upload_dir, filename)
    os.makedirs(upload_dir, exist_ok=True)
    with open(filepath, "wb") as f:
        f.write(content)
    return filepath
